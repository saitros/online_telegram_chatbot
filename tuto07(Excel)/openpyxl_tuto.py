from openpyxl import load_workbook

# 사용할 엑셀 파일 경로
EXCEL_FILE_NAME = 'Database.xlsx'

# 엑셀 파일 객체
db = load_workbook(filename=EXCEL_FILE_NAME)

# 엑셀 웨크 시트 객체
tuto_db = db['Tuto']
user_db = db['User']


def write_with_index():
    """
    엑셀의 index 를 통해서 셀 접근 방법
    .value 를 넣어줘야 해당 셀의 값을 확인 가능
    .value 가 없으면 해당 셀의 위치정보 확인
    """
    tuto_db['A1'].value = 100
    print(tuto_db['A1'])
    print(tuto_db['A1'].value)


def write_with_cell():
    """
    엑셀의 Cell 정보를 통해서 접근하는 방법
    .cell의 매개변수로 row, column 값을 (1~ ) 전달
    값을 수정하려면 value 매개변수로 인자를 전달
    값을 읽을때 마찬가지로 .value 를 해서 값을 읽기
    """
    tuto_db.cell(row=1, column=1, value='안녕하세요')
    print(tuto_db.cell(row=1, column=1))
    print(tuto_db.cell(row=1, column=1).value)


def db_user_set(user_id, name, msg):
    user_exist = False

    for row in user_db.rows:
        if row[0].value == user_id:
            user_exist = True
            user_prev_msg = row[2].value
            row[2].value = msg

    if not user_exist:
        user_prev_msg = None
        user_db[user_db.max_row + 1][0].value = user_id
        user_db[user_db.max_row][1].value = name
        user_db[user_db.max_row][2].value = msg

    db.save(EXCEL_FILE_NAME)
    return user_prev_msg


if __name__ == '__main__':
    # write_with_index()
    write_with_cell()

    # 엑셀파일을 저장
    # 경로를 수정하면 다른이름으로 저장 가능 없는 파일일 경우 파일 생성
    db.save(EXCEL_FILE_NAME)
    print('ok')
